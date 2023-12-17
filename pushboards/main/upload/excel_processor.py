from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from pushboards.main.upload.openpyxl_get_colors import OpenpyxlColorToRgbaConverter, Theme

CONF_SHEET_NAME = "Настройка импорта"


def get_import_configuration(conf_sheet: Worksheet, get_cell_color: OpenpyxlColorToRgbaConverter) -> dict[str, str]:
    conf = {}
    for row in conf_sheet:
        for cell in row:
            cell_color = get_cell_color(cell.fill.fgColor)
            if cell_color:
                conf[cell_color] = str(cell.value or "")
    return conf


def import_raw_data(
    data_sheet: Worksheet, get_cell_color: OpenpyxlColorToRgbaConverter, conf: dict[str, str]
) -> list[list[str]]:
    data = []
    for col_id, col in data_sheet.column_dimensions.items():
        column_color = get_cell_color(col.fill.fgColor)
        if column_color not in conf:
            continue
        values = [conf[column_color]]
        for row_id in data_sheet.row_dimensions.keys():
            values.append(data_sheet[f"{col_id}{row_id}"].value or "")
        data.append(values)
    return data


def process1(file_data: Path | BytesIO) -> pd.DataFrame:
    # read excel file
    workbook = load_workbook(file_data)

    if CONF_SHEET_NAME not in workbook.sheetnames:
        raise ValueError("Configuration sheet not found")

    # function to reliably read cells' colors
    get_cell_color = OpenpyxlColorToRgbaConverter(Theme(workbook))

    # get excel sheets
    data_sheet = workbook[workbook.sheetnames[0]]
    conf_sheet = workbook[CONF_SHEET_NAME]

    # read configuration from the configuration sheet
    conf = get_import_configuration(conf_sheet, get_cell_color)

    if not conf:
        raise ValueError("Configuration sheet is empty")

    # read data from the data sheet
    data = import_raw_data(data_sheet, get_cell_color, conf)

    if not data:
        raise ValueError("Data sheet is empty")

    # convert the raw data to a Pandas DataFrame
    result = pd.DataFrame(data).transpose()
    new_index = list(result.index)
    new_index[:2] = (1, 0)
    result = result.reindex(new_index)
    result = result.replace(r"^\s*$", None, regex=True).ffill(axis=1).ffill(axis=0).fillna("")

    return result
