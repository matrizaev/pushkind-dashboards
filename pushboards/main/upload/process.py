from io import BytesIO
from pathlib import Path
from typing import Any, Callable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from pushboards.main.upload.openpyxl_get_colors import OpenpyxlColorToRgbaConverter, Theme

ProcessCallable = Callable[[bytes], bytes]


def update_value_list_dict(data: dict, key: Any, value: Any) -> None:
    # Find the maximum length of lists in the dictionary values
    max_len = max(map(len, data.values()), default=0)

    # If the key is not already in the dictionary, initialize it with an empty list
    data.setdefault(key, [])

    # Ensure all lists in the dictionary have the same length
    for k, v in data.items():
        len_diff = max_len - len(v) - 1
        if len_diff > 0:
            data[k] += [""] * len_diff

    data[key].append(value)


def get_templated_data(data_sheet, template_sheet):
    template_cells = []
    for row in template_sheet.iter_rows():
        row_contains_data = False
        for cell in row:
            if cell.fill.fgColor.rgb in ("FFFFFFFF", "00000000"):
                continue
            if not row_contains_data:
                row_contains_data = True
                template_cells.append({})
            if not cell.value:
                continue
            keys = cell.value.split(":", maxsplit=1)
            if len(keys) == 1:
                template_cells[-1][keys[0]] = data_sheet[cell.coordinate].value
            else:
                if keys[0] not in template_cells[-1]:
                    template_cells[-1][keys[0]] = {}
                update_value_list_dict(template_cells[-1][keys[0]], keys[1], data_sheet[cell.coordinate].value)
    return template_cells


def update_output_sheet(output_sheet, template_cells, header_row=3):
    header = [cell.value for cell in output_sheet[header_row]]
    start_row = output_sheet.max_row + 1
    start_col = 1

    for entity in template_cells:
        for key, value in entity.items():
            try:
                col_idx = header.index(key) + 1
            except ValueError:
                continue
            if not isinstance(value, dict):
                cell = output_sheet.cell(start_row, col_idx)
                cell.value = value
                continue

            for subkey, subvalues in value.items():
                cell = output_sheet.cell(start_row, col_idx)
                cell.value = subkey
                start_col = col_idx + 1
                for subvalue in subvalues:
                    cell = output_sheet.cell(start_row, start_col)
                    cell.value = subvalue
                    start_col += 1
                start_row += 1
        start_row += 1


def get_import_configuration(conf_sheet: Worksheet, get_cell_color: OpenpyxlColorToRgbaConverter) -> dict[str, str]:
    conf = {}
    for row in conf_sheet:
        for cell in row:
            cell_color = get_cell_color(cell.fill.fgColor)
            if cell_color:
                conf[cell_color] = str(cell.value or "")
    return conf


def import_raw_data(
    data_sheet: Worksheet, get_cell_color: OpenpyxlColorToRgbaConverter, conf: dict[str, str]
) -> list[list[str]]:
    data = []
    for col_id, col in data_sheet.column_dimensions.items():
        column_color = get_cell_color(col.fill.fgColor)
        if column_color not in conf:
            continue
        values = [conf[column_color]]
        for row_id in data_sheet.row_dimensions.keys():
            values.append(data_sheet[f"{col_id}{row_id}"].value or "")
        data.append(values)
    return data


def process(file_data: BytesIO, conf_sheet_name: str) -> BytesIO:
    # read excel file
    workbook = load_workbook(file_data)

    if conf_sheet_name not in workbook.sheetnames:
        raise ValueError("Configuration sheet not found")

    # function to reliably read cells' colors
    get_cell_color = OpenpyxlColorToRgbaConverter(Theme(workbook))

    # get excel sheets
    data_sheet = workbook[workbook.sheetnames[0]]
    conf_sheet = workbook[conf_sheet_name]

    # read configuration from the configuration sheet
    conf = get_import_configuration(conf_sheet, get_cell_color)

    if not conf:
        raise ValueError("Configuration sheet is empty")

    # read data from the data sheet
    data = import_raw_data(data_sheet, get_cell_color, conf)

    if not data:
        raise ValueError("Data sheet is empty")

    # convert the raw data to a Pandas DataFrame
    result = pd.DataFrame(data).transpose().fillna("")

    # save to a BytesIO
    result_data = BytesIO()
    result.to_pickle(result_data)
    result_data.seek(0)
    return result_data


def pandas_pickle_to_excel(path: Path | BytesIO) -> BytesIO:
    # read from pickle
    result = pd.read_pickle(path)

    # save to excel
    result_data = BytesIO()
    result.to_excel(result_data, index=False, header=False)
    result_data.seek(0)
    return result_data


def pandas_pickle_to_html(file_path: Path | BytesIO) -> str:
    # read from pickle
    result = pd.read_pickle(file_path)

    # save to html
    return result.to_html(index=False, header=False, justify="left", classes=["table", "table-striped"])
