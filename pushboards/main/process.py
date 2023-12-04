from io import BytesIO
from typing import Any

from openpyxl import load_workbook


def update_value_list_dict(data: dict, key: Any, value: Any) -> None:
    # Find the maximum length of lists in the dictionary values
    max_len = max(map(len, data.values()), default=0)

    # If the key is not already in the dictionary, initialize it with an empty list
    data.setdefault(key, [])

    # Ensure all lists in the dictionary have the same length
    for k, v in data.items():
        len_diff = max_len - len(v) - 1
        if len_diff > 0:
            data[k] += [""] * len_diff

    data[key].append(value)


def get_templated_data(data_sheet, template_sheet):
    template_cells = []
    for row in template_sheet.iter_rows():
        row_contains_data = False
        for cell in row:
            if cell.fill.fgColor.rgb in ("FFFFFFFF", "00000000"):
                continue
            if not row_contains_data:
                row_contains_data = True
                template_cells.append({})
            if not cell.value:
                continue
            keys = cell.value.split(":", maxsplit=1)
            if len(keys) == 1:
                template_cells[-1][keys[0]] = data_sheet[cell.coordinate].value
            else:
                if keys[0] not in template_cells[-1]:
                    template_cells[-1][keys[0]] = {}
                update_value_list_dict(template_cells[-1][keys[0]], keys[1], data_sheet[cell.coordinate].value)
    return template_cells


def update_output_sheet(output_sheet, template_cells, header_row=3):
    header = [cell.value for cell in output_sheet[header_row]]
    start_row = output_sheet.max_row + 1
    start_col = 1

    for entity in template_cells:
        for key, value in entity.items():
            try:
                col_idx = header.index(key) + 1
            except ValueError:
                continue
            if not isinstance(value, dict):
                cell = output_sheet.cell(start_row, col_idx)
                cell.value = value
                continue

            for subkey, subvalues in value.items():
                cell = output_sheet.cell(start_row, col_idx)
                cell.value = subkey
                start_col = col_idx + 1
                for subvalue in subvalues:
                    cell = output_sheet.cell(start_row, start_col)
                    cell.value = subvalue
                    start_col += 1
                start_row += 1
        start_row += 1


def process(file_data: bytes) -> bytes:
    # read excel file
    workbook = load_workbook(file_data)

    # get excel sheets
    template_sheet = workbook[workbook.sheetnames[2]]
    data_sheet = workbook[workbook.sheetnames[0]]
    output_sheet = workbook[workbook.sheetnames[1]]

    template_cells = get_templated_data(data_sheet, template_sheet)
    update_output_sheet(output_sheet, template_cells)

    result_data = BytesIO()
    workbook.save(result_data)
    return result_data.getvalue()
